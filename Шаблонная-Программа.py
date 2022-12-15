import PySimpleGUI as sg
import openpyxl as xl
import os
from docxtpl import DocxTemplate
import warnings

warnings.simplefilter("ignore")

Groups = ['']
Mans = ['']
Job = ['']
SelJob = ['']
SelMans = ['']
card_content = []
# Сдесь реализованно очень простое подключение шаблона
# Для начала я выбираю путь к вашему рабочему столу
path = os.path.expanduser('~\\Desktop')
os.chdir(path)
# Далее ищу файл с заполенным шаблоном
template = DocxTemplate("Шаблон.docx")

# Графика
layout = [
    [sg.Text('Файл, со значениями 1'), sg.InputText(k='-job-'), sg.FileBrowse(k='-job-')],
    [sg.Text('Файл, со значениями 2'), sg.InputText(k='-man-'), sg.FileBrowse(k='-man-')],
    [sg.Text('Сортируете по этому значению'), sg.DD(Groups, size=(52, 8), default_value=Groups[0], k='-Group-')],
    [sg.Text('Выбор значения 1'), sg.DD(Mans, size=(52, 8), default_value='', k='-Name-')],
    [sg.Text('Выбор значения 2'), sg.DD(Job, size=(39, 8), default_value='', k='-Job-')],
    [sg.Output(size=(72, 20), k='-Terminal-')],
    [sg.Push(), sg.Btn('Fill'), sg.Btn('Next'), sg.Btn('Connect'), sg.Btn('Clear'), sg.Cancel(), sg.Push()]
]

# Имя окна
window = sg.Window('Тут можно обозвать вашу программу', layout)

while True:
    # Чтение переменных с окна
    event, values = window.read()

    # Программирование кнопок

    # Программирование закрытия программы
    if event in ('Exit', 'Cancel', sg.WINDOW_CLOSED):
        break

    # кнопка для очистки терминала
    if event == 'Clear':
        window['-Terminal-'].update('')

    if event == 'Connect':
        if values['-job-'] is None or values['-job-'] == '':
            print("Вы не выбрали путь к файлу 1!")
            continue
        elif values['-man-'] is None or values['-man-'] == '':
            print("Вы не выбрали путь к файлу 2!")
            continue
        # Подключение к файлу где храняться дорожные файлы
        workbook = xl.load_workbook(values['-job-'])
        sheet_1 = workbook.active
        # Подключение к файлу с именами студентов
        workbook2 = xl.load_workbook(values['-man-'])
        sheet_2 = workbook2.active
        # Создаем массив с группами и наименованиями дисциплин/МДК/ПМ
        tJob = []
        tGroups = [""]
        for i in range(1, sheet_1.max_row):
            tJob.append(sheet_1.cell(i, 8).value)
            tGroups.append(sheet_1.cell(i, 5).value)

        # Аналогично сделаем с именами студентов
        tMans = []
        for i in range(1, sheet_2.max_row):
            tMans.append(sheet_2.cell(i, 6).value)

        # Тут мы подчистим массивы от лишних и повторяющихся значений,
        # если вы начинаете читать файл с первой строки, добавьте исключение наименование сстолюбца

        # Для Job
        for el in tJob:
            if el not in Job and el is not None and el != '':
                Job.append(el)

        # Для Groups
        for el in tGroups:
            if el not in Groups and el is not None and el != '':
                Groups.append(el)
        Groups.sort()

        window['-Group-'].update(Groups[0], Groups)
        print("Вы успешно подключили файлы!")

    # Сортировка по выбранному параметру
    if event == 'Next':
        if values['-Group-'] == '':
            print("Вы забыли ввести группу!")
            continue
        a = 0
        b = 0
        # Тут я как раз сортирую значение по выбранному параметру
        for i in range(2, sheet_1.max_row + 1):
            if sheet_1.cell(i, 5).value == values['-Group-']:
                SelJob.append(sheet_1.cell(i, 8).value)
            for el in SelJob:
                if el not in SelJob and el != None and el != '':
                    SelJob.append(el)
        SelJob.sort()
        for i in range(1, sheet_2.max_row + 1):
            a = str(values['-Group-']).replace(" ", "")
            b = str(sheet_2.cell(i, 1).value).replace(" ", "")
            if b == a:
                SelMans.append(sheet_2.cell(i, 6).value)
            for el in SelMans:
                if el not in SelMans and el != None and el != '':
                    SelMans.append(el)
        SelMans.sort()

        window['-Name-'].update('', SelMans)
        window['-MDK-'].update('', SelJob)
        print("Теперь выбирите Человека и Работу")

    # Тут идет заполнение шаблона
    if event == 'Fill':
        if values['-Name-'] == '':
            print("Вы забыли выбрать Человека!")
            continue
        if values['-MDK-'] == '':
            print("Вы забыли выбрать Работу!")
            continue
        # Заполнение шаблона выбранными данными
        for i in range(2, sheet_1.max_row + 1):
            # проверка на наличие нужных данных и их ввод в шаблон
            if sheet_1.cell(i, 5).value == values['-Group-'] and sheet_1.cell(i, 8).value == values['-Job-']:
                card_content = {
                    'Mans': values['-Name-'],
                    'job': sheet_1.cell(i, 8).value,
                    'group': sheet_1.cell(i, 1).value
                }
            else:
                continue

        # сохранение итога в отдельный файл
        template.render(card_content)
        template.save('Заполненный Шаблон.docx')

window.close()
